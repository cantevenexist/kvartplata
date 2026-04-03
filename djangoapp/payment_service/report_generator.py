"""
Простой генератор Excel отчетов
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.core.files.base import ContentFile
from decimal import Decimal
from datetime import datetime
from io import BytesIO

from .models import Report, Charge, Payment, Debt
from housing.models import HousingUnit


def generate_full_report(created_by, period_start=None, period_end=None):
    """
    Генерация полного отчета с тремя листами:
    - Начисления
    - Платежи
    - Задолженности
    """
    wb = openpyxl.Workbook()
    
    # Удаляем стандартный лист
    wb.remove(wb.active)
    
    # Создаем три листа
    ws_charges = wb.create_sheet("Начисления")
    ws_payments = wb.create_sheet("Платежи")
    ws_debts = wb.create_sheet("Задолженности")
    
    # Заполняем каждый лист
    fill_charges_sheet(ws_charges, period_start, period_end)
    fill_payments_sheet(ws_payments, period_start, period_end)
    fill_debts_sheet(ws_debts, period_start, period_end)
    
    # Сохраняем файл
    filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    report = Report(
        report_type='full',
        created_by=created_by,
        period_start=period_start,
        period_end=period_end,
    )
    report.file.save(filename, ContentFile(output.getvalue()), save=True)
    
    return report


def fill_charges_sheet(ws, period_start=None, period_end=None):
    """Заполняет лист с начислениями"""
    
    # Заголовки
    headers = ["ID", "ID жилья", "Адрес", "Владелец", "Период", "Тариф", "Сумма", "Оплачено", "Остаток", "Статус"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Получаем данные
    charges = Charge.objects.select_related('tariff').order_by('-period')
    
    if period_start and period_end:
        charges = charges.filter(period__gte=period_start, period__lte=period_end)
    
    # Заполняем данные
    row = 2
    for charge in charges:
        # Получаем адрес и владельца
        address = f"Жилье #{charge.housing_id}"
        owner_name = "—"
        try:
            unit = HousingUnit.objects.get(id=charge.housing_id)
            address = unit.address
            if unit.owner:
                try:
                    owner_name = unit.owner.profile.full_name or unit.owner.username
                except:
                    owner_name = unit.owner.username
        except:
            pass
        
        remaining = charge.amount - charge.paid_amount
        status = "Оплачено" if charge.is_paid else ("Частично" if charge.paid_amount > 0 else "Не оплачено")
        
        ws.cell(row=row, column=1, value=charge.id)
        ws.cell(row=row, column=2, value=charge.housing_id)
        ws.cell(row=row, column=3, value=address)
        ws.cell(row=row, column=4, value=owner_name)
        ws.cell(row=row, column=5, value=charge.period.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=6, value=f"{charge.tariff.name} ({charge.tariff.rate_per_unit} руб./{charge.tariff.get_unit_display()})")
        ws.cell(row=row, column=7, value=float(charge.amount))
        ws.cell(row=row, column=8, value=float(charge.paid_amount))
        ws.cell(row=row, column=9, value=float(remaining))
        ws.cell(row=row, column=10, value=status)
        row += 1
    
    # Автоширина
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18


def fill_payments_sheet(ws, period_start=None, period_end=None):
    """Заполняет лист с платежами"""
    
    # Заголовки
    headers = ["ID", "ID жилья", "Адрес", "Владелец", "Сумма", "Способ оплаты", "Дата платежа", "Описание"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Получаем данные
    payments = Payment.objects.order_by('-payment_date')
    
    if period_start and period_end:
        payments = payments.filter(payment_date__date__gte=period_start, payment_date__date__lte=period_end)
    
    # Заполняем данные
    row = 2
    for payment in payments:
        # Получаем адрес и владельца
        address = f"Жилье #{payment.housing_id}"
        owner_name = "—"
        try:
            unit = HousingUnit.objects.get(id=payment.housing_id)
            address = unit.address
            if unit.owner:
                try:
                    owner_name = unit.owner.profile.full_name or unit.owner.username
                except:
                    owner_name = unit.owner.username
        except:
            pass
        
        ws.cell(row=row, column=1, value=payment.id)
        ws.cell(row=row, column=2, value=payment.housing_id)
        ws.cell(row=row, column=3, value=address)
        ws.cell(row=row, column=4, value=owner_name)
        ws.cell(row=row, column=5, value=float(payment.amount))
        ws.cell(row=row, column=6, value=payment.get_payment_method_display())
        ws.cell(row=row, column=7, value=payment.payment_date.strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=row, column=8, value=payment.description or "")
        row += 1
    
    # Автоширина
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18


def fill_debts_sheet(ws, period_start=None, period_end=None):
    """Заполняет лист с задолженностями"""
    
    # Заголовки
    headers = ["ID", "ID жилья", "Адрес", "Владелец", "Период", "Сумма долга"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Получаем данные
    debts = Debt.objects.filter(total_amount__gt=0).order_by('-period')
    
    if period_start and period_end:
        debts = debts.filter(created_at__date__gte=period_start, created_at__date__lte=period_end)
    
    # Заполняем данные
    row = 2
    total_debt = Decimal('0')
    
    for debt in debts:
        # Получаем адрес и владельца
        address = f"Жилье #{debt.housing_id}"
        owner_name = "—"
        try:
            unit = HousingUnit.objects.get(id=debt.housing_id)
            address = unit.address
            if unit.owner:
                try:
                    owner_name = unit.owner.profile.full_name or unit.owner.username
                except:
                    owner_name = unit.owner.username
        except:
            pass
        
        ws.cell(row=row, column=1, value=debt.id)
        ws.cell(row=row, column=2, value=debt.housing_id)
        ws.cell(row=row, column=3, value=address)
        ws.cell(row=row, column=4, value=owner_name)
        ws.cell(row=row, column=5, value=debt.period.strftime("%B %Y") if debt.period else "—")
        ws.cell(row=row, column=6, value=float(debt.total_amount))
        
        total_debt += debt.total_amount
        row += 1
    
    # Итог
    if row > 2:
        ws.cell(row=row, column=5, value="ИТОГО:")
        ws.cell(row=row, column=6, value=float(total_debt))
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6).font = Font(bold=True)
    
    # Автоширина
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18